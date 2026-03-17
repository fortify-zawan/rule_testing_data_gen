"""Export a test suite to CSV, JSON, or XLSX."""
import csv
import io
import json

from domain.models import BehavioralTestCase, Rule, Transaction

# ─── CSV ──────────────────────────────────────────────────────────────────────

def export_csv(
    rule: Rule,
    sequence: list[Transaction] | None,
    cases: list[BehavioralTestCase],
) -> str:
    output = io.StringIO()
    writer = csv.writer(output)

    if rule.rule_type == "stateless" and sequence:
        # One row per tagged transaction
        headers = ["id", "tag"] + rule.relevant_attributes + ["expected", "validation"]
        writer.writerow(headers)
        for t in sequence:
            row = [t.id, t.tag] + [t.attributes.get(a, "") for a in rule.relevant_attributes]
            if t.validation_result:
                row += [
                    "trigger" if t.validation_result.expected_trigger else "no_trigger",
                    "PASS" if t.validation_result.passed else "FAIL",
                ]
            else:
                row += ["—", "—"]
            writer.writerow(row)

    elif rule.rule_type == "behavioral" and cases:
        # Summary: one row per test case
        agg_keys = list(cases[0].computed_aggregates.keys()) if cases else []
        headers = ["tc_id", "type", "n_transactions", "expected", "validation"] + agg_keys
        writer.writerow(headers)
        for case in cases:
            vr = case.validation_result
            row = [
                case.id,
                case.scenario_type,
                len(case.transactions),
                "trigger" if (vr and vr.expected_trigger) else "no_trigger",
                "PASS" if (vr and vr.passed) else "FAIL",
            ] + [case.computed_aggregates.get(k, "") for k in agg_keys]
            writer.writerow(row)

    return output.getvalue()


# ─── JSON ─────────────────────────────────────────────────────────────────────

def _transaction_to_dict(t: Transaction) -> dict:
    d = {"id": t.id, "tag": t.tag, "attributes": t.attributes}
    if t.validation_result:
        vr = t.validation_result
        d["validation"] = {
            "passed": vr.passed,
            "expected_trigger": vr.expected_trigger,
            "conditions": [
                {
                    "attribute": cr.attribute,
                    "operator": cr.operator,
                    "threshold": cr.threshold,
                    "actual": cr.actual_value,
                    "passed": cr.passed,
                }
                for cr in vr.condition_results
            ],
        }
    return d


def export_json(
    rule: Rule,
    sequence: list[Transaction] | None,
    cases: list[BehavioralTestCase],
) -> str:
    payload = {
        "rule": {
            "description": rule.description,
            "rule_type": rule.rule_type,
            "raw_expression": rule.raw_expression,
            "conditions": [
                {
                    "attribute": c.attribute,
                    "operator": c.operator,
                    "value": c.value,
                    "aggregation": c.aggregation,
                    "logical_connector": c.logical_connector,
                }
                for c in rule.conditions
            ],
        }
    }

    if rule.rule_type == "stateless" and sequence:
        payload["sequence"] = [_transaction_to_dict(t) for t in sequence]

    elif rule.rule_type == "behavioral" and cases:
        payload["test_cases"] = [
            {
                "id": case.id,
                "scenario_type": case.scenario_type,
                "intent": case.intent,
                "transactions": [_transaction_to_dict(t) for t in case.transactions],
                "computed_aggregates": case.computed_aggregates,
                "validation": {
                    "passed": case.validation_result.passed,
                    "expected_trigger": case.validation_result.expected_trigger,
                    "conditions": [
                        {
                            "attribute": cr.attribute,
                            "operator": cr.operator,
                            "threshold": cr.threshold,
                            "actual": cr.actual_value,
                            "passed": cr.passed,
                        }
                        for cr in case.validation_result.condition_results
                    ],
                } if case.validation_result else None,
            }
            for case in cases
        ]

    return json.dumps(payload, indent=2, default=str)


# ─── XLSX ─────────────────────────────────────────────────────────────────────

def export_xlsx(
    rule: Rule,
    sequence: list[Transaction] | None,
    cases: list[BehavioralTestCase],
) -> bytes:
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        # Fallback: return CSV bytes with xlsx extension
        return export_csv(rule, sequence, cases).encode()

    wb = openpyxl.Workbook()

    green_fill = PatternFill("solid", fgColor="C6EFCE")
    red_fill = PatternFill("solid", fgColor="FFC7CE")
    header_font = Font(bold=True)

    if rule.rule_type == "stateless" and sequence:
        ws = wb.active
        ws.title = "Stateless Sequence"
        headers = ["id", "tag"] + rule.relevant_attributes + ["expected", "validation"]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = header_font

        for t in sequence:
            row = [t.id, t.tag] + [t.attributes.get(a, "") for a in rule.relevant_attributes]
            if t.validation_result:
                row += [
                    "trigger" if t.validation_result.expected_trigger else "no_trigger",
                    "PASS" if t.validation_result.passed else "FAIL",
                ]
            else:
                row += ["—", "—"]
            ws.append(row)
            # Colour validation cell
            val_col = len(headers)
            cell = ws.cell(ws.max_row, val_col)
            if cell.value == "PASS":
                cell.fill = green_fill
            elif cell.value == "FAIL":
                cell.fill = red_fill

    elif rule.rule_type == "behavioral" and cases:
        ws = wb.active
        ws.title = "Test Cases Summary"
        agg_keys = list(cases[0].computed_aggregates.keys()) if cases else []
        headers = ["tc_id", "type", "n_transactions", "expected", "validation"] + agg_keys
        ws.append(headers)
        for cell in ws[1]:
            cell.font = header_font

        for case in cases:
            vr = case.validation_result
            row = [
                case.id,
                case.scenario_type,
                len(case.transactions),
                "trigger" if (vr and vr.expected_trigger) else "no_trigger",
                "PASS" if (vr and vr.passed) else "FAIL",
            ] + [case.computed_aggregates.get(k, "") for k in agg_keys]
            ws.append(row)
            cell = ws.cell(ws.max_row, 5)
            if cell.value == "PASS":
                cell.fill = green_fill
            elif cell.value == "FAIL":
                cell.fill = red_fill

        # One sheet per test case with transaction detail
        for i, case in enumerate(cases):
            ws2 = wb.create_sheet(title=f"TC-{i+1}-{case.scenario_type[:3]}")
            all_attrs = list({k for t in case.transactions for k in t.attributes})
            row_headers = ["id", "tag"] + all_attrs
            ws2.append(row_headers)
            for cell in ws2[1]:
                cell.font = header_font
            for t in case.transactions:
                ws2.append([t.id, t.tag] + [t.attributes.get(a, "") for a in all_attrs])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
